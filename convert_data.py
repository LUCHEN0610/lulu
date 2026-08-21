#!/usr/bin/env python3
"""
数据转换脚本 - 从Excel文件生成JSON数据
使用方法: python3 convert_data.py
依赖: pip install openpyxl
"""

import openpyxl
import json
import os
import sys

def convert_category_guide(excel_path, output_path):
    """转换类目指引.xlsx → category_guide.json"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Sheet1']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
    
    data = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers):
            val = ws.cell(r, c + 1).value
            if val is not None:
                row[h] = str(val).strip()
        if row:
            data.append(row)
    
    # 按类目分组，合并相同类目分类的路径
    result = {}
    for item in data:
        cat = item.get('类目', '').strip()
        sub = item.get('类目分类', '').strip()
        path = item.get('类目路径', '').strip()
        if not cat:
            continue
        if cat not in result:
            result[cat] = {}
        if sub not in result[cat]:
            result[cat][sub] = []
        if path and path not in result[cat][sub]:
            result[cat][sub].append(path)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"✅ 类目指引: {output_path} ({len(result)} 个类目)")


def convert_market_demand(excel_path, output_path):
    """转换招品.xlsx → market_demand.json"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['任务数据']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    
    data = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers):
            val = ws.cell(r, c + 1).value
            row[h] = str(val).strip() if val is not None else ''
        
        if not row.get('行业图片url'):
            continue
        
        # 解析图片URL
        img_urls = row['行业图片url']
        if '","' in img_urls:
            img_urls = img_urls.replace('"', '')
            urls = [u.strip() for u in img_urls.split(',') if u.strip()]
        elif img_urls.startswith('['):
            try:
                urls = json.loads(img_urls)
            except:
                urls = [img_urls]
        else:
            urls = [img_urls]
        
        row['images'] = urls
        row['image'] = urls[0] if urls else ''
        row['region'] = row.get('区域', '').strip()
        row['category'] = row.get('类目', '').strip()
        
        data.append(row)
    
    # 按更新时间倒序
    data.sort(key=lambda x: x.get('提需时间', ''), reverse=True)
    
    result = {
        'items': data,
        'regions': ['欧区', '美区', '拉美'],
        'categories': ['外套夹克', '衬衫']
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"✅ 市场需求: {output_path} ({len(data)} 条记录)")


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, 'data')
    
    # 确保data目录存在
    os.makedirs(data_dir, exist_ok=True)
    
    # 查找Excel文件
    excel_files = {
        'category': None,
        'demand': None
    }
    
    # 在当前目录和上级目录查找
    for search_dir in ['.', '..', os.path.join('..', '..')]:
        for f in os.listdir(search_dir):
            if '类目指引' in f and f.endswith('.xlsx'):
                excel_files['category'] = os.path.join(search_dir, f)
            if '招品' in f and f.endswith('.xlsx'):
                excel_files['demand'] = os.path.join(search_dir, f)
    
    if excel_files['category']:
        convert_category_guide(excel_files['category'], os.path.join(data_dir, 'category_guide.json'))
    else:
        print("⚠️ 未找到类目指引.xlsx，跳过")
    
    if excel_files['demand']:
        convert_market_demand(excel_files['demand'], os.path.join(data_dir, 'market_demand.json'))
    else:
        print("⚠️ 未找到招品.xlsx，跳过")
    
    print("\n📊 数据转换完成！")


if __name__ == '__main__':
    main()
